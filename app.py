import io
import os
import sqlite3
from datetime import datetime
from functools import wraps

import pandas as pd
from flask import (Flask, abort, flash, jsonify, redirect, render_template, request,
                   send_file, send_from_directory, session, url_for)
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'change-this-in-prod-a8f3k2p9')


# Register Jinja filters
@app.template_filter('format_description')
def filter_format_description(text):
    """Render formatted description as HTML."""
    from markupsafe import Markup
    if not text or not text.strip():
        return ''

    sections = format_description(text)
    html_parts = []

    for section in sections:
        if section['type'] == 'header':
            html_parts.append(f'<div class="desc-header">{section["content"]}</div>')
        elif section['type'].startswith('h'):
            level = section['type'][1]
            html_parts.append(f'<h{level} class="desc-heading">{section["content"]}</h{level}>')
        elif section['type'] == 'paragraph':
            html_parts.append(f'<p class="desc-paragraph">{section["content"]}</p>')
        elif section['type'] == 'list':
            items_html = ''.join(f'<li>{item}</li>' for item in section['items'])
            html_parts.append(f'<ul class="desc-list">{items_html}</ul>')
        elif section['type'] == 'spacing':
            pass  # Skip spacing in rendered output

    return Markup('\n'.join(html_parts))


# ── Context Processors ────────────────────────────────────────────────────────
@app.context_processor
def inject_projects():
    """Make all projects and active project available to all templates."""
    projects = []
    active_project_name = None
    if 'user_id' in session:
        with get_db() as conn:
            projects = conn.execute(
                'SELECT id, name FROM projects ORDER BY name'
            ).fetchall()
            if session.get('active_project_id'):
                active_proj = conn.execute(
                    'SELECT name FROM projects WHERE id = ?',
                    (session.get('active_project_id'),)
                ).fetchone()
                active_project_name = active_proj['name'] if active_proj else None
    return {
        'all_projects': projects,
        'active_project_name': active_project_name
    }

# ── Storage paths ────────────────────────────────────────────────────────────
DATA_DIR = os.environ.get('DATA_DIR', os.path.join(os.path.dirname(__file__), 'data'))
DB_PATH = os.path.join(DATA_DIR, 'procurement.db')
UPLOAD_FOLDER = os.path.join(DATA_DIR, 'uploads')
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

ALLOWED_EXT = {'xlsx', 'xls', 'csv'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


# ── Database ─────────────────────────────────────────────────────────────────
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def save_mapping(mapping_type, mapping_dict):
    """Persist the last-used column mapping for a given type (items / quotes)."""
    import json as _json
    with get_db() as conn:
        conn.execute('''
            INSERT INTO column_mappings (mapping_type, mapping_json)
            VALUES (?, ?)
            ON CONFLICT(mapping_type) DO UPDATE SET mapping_json = excluded.mapping_json
        ''', (mapping_type, _json.dumps(mapping_dict)))


def load_mapping(mapping_type):
    """Load the last-used column mapping for a given type, or empty dict."""
    import json as _json
    with get_db() as conn:
        row = conn.execute(
            'SELECT mapping_json FROM column_mappings WHERE mapping_type = ?',
            (mapping_type,)
        ).fetchone()
    return _json.loads(row['mapping_json']) if row else {}


def init_db():
    with get_db() as conn:
        conn.executescript('''
            CREATE TABLE IF NOT EXISTS users (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                username      TEXT UNIQUE NOT NULL,
                password_hash TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS items (
                id                    INTEGER PRIMARY KEY AUTOINCREMENT,
                name                  TEXT NOT NULL,
                category              TEXT,
                description           TEXT,
                supplier_description  TEXT,
                qty                   REAL,
                unit                  TEXT,
                created_at            TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS suppliers (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                name       TEXT NOT NULL,
                phone      TEXT,
                address    TEXT,
                remarks    TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS quote_requests (
                id         INTEGER PRIMARY KEY AUTOINCREMENT,
                title      TEXT NOT NULL,
                notes      TEXT,
                status     TEXT DEFAULT 'open',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS quote_request_items (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                quote_request_id INTEGER NOT NULL,
                item_id          INTEGER NOT NULL,
                FOREIGN KEY (quote_request_id) REFERENCES quote_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (item_id)          REFERENCES items(id)          ON DELETE CASCADE,
                UNIQUE(quote_request_id, item_id)
            );

            CREATE TABLE IF NOT EXISTS quote_request_suppliers (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                quote_request_id INTEGER NOT NULL,
                supplier_id      INTEGER NOT NULL,
                FOREIGN KEY (quote_request_id) REFERENCES quote_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (supplier_id)      REFERENCES suppliers(id)      ON DELETE CASCADE,
                UNIQUE(quote_request_id, supplier_id)
            );

            CREATE TABLE IF NOT EXISTS supplier_quotes (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                quote_request_id INTEGER NOT NULL,
                supplier_id      INTEGER NOT NULL,
                file_path        TEXT,
                notes            TEXT,
                uploaded_at      TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (quote_request_id) REFERENCES quote_requests(id) ON DELETE CASCADE,
                FOREIGN KEY (supplier_id)      REFERENCES suppliers(id)      ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS quote_prices (
                id               INTEGER PRIMARY KEY AUTOINCREMENT,
                supplier_quote_id INTEGER NOT NULL,
                item_id          INTEGER NOT NULL,
                price            REAL,
                notes            TEXT,
                FOREIGN KEY (supplier_quote_id) REFERENCES supplier_quotes(id) ON DELETE CASCADE,
                FOREIGN KEY (item_id)           REFERENCES items(id)           ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS column_mappings (
                mapping_type TEXT PRIMARY KEY,
                mapping_json TEXT NOT NULL
            );

            CREATE TABLE IF NOT EXISTS projects (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                name        TEXT NOT NULL,
                description TEXT,
                status      TEXT DEFAULT 'active',
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            );

            CREATE TABLE IF NOT EXISTS item_lists (
                id          INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id  INTEGER NOT NULL,
                name        TEXT NOT NULL,
                description TEXT,
                created_at  TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (project_id) REFERENCES projects(id) ON DELETE CASCADE
            );

            CREATE TABLE IF NOT EXISTS item_list_items (
                id           INTEGER PRIMARY KEY AUTOINCREMENT,
                item_list_id INTEGER NOT NULL,
                item_id      INTEGER NOT NULL,
                FOREIGN KEY (item_list_id) REFERENCES item_lists(id) ON DELETE CASCADE,
                FOREIGN KEY (item_id)      REFERENCES items(id)      ON DELETE CASCADE,
                UNIQUE(item_list_id, item_id)
            );
        ''')

        # Migrate: add supplier_description column if it doesn't exist (for existing DBs)
        try:
            conn.execute('ALTER TABLE items ADD COLUMN supplier_description TEXT')
        except Exception:
            pass  # Column already exists

        # Migrate: add item_list_id column to quote_requests (for existing DBs)
        try:
            conn.execute('ALTER TABLE quote_requests ADD COLUMN item_list_id INTEGER REFERENCES item_lists(id)')
        except Exception:
            pass  # Column already exists

        # Seed default admin account
        admin_username = os.environ.get('ADMIN_USERNAME', 'admin')
        admin_password = os.environ.get('ADMIN_PASSWORD', 'admin123')
        row = conn.execute('SELECT id FROM users WHERE username = ?', (admin_username,)).fetchone()
        if not row:
            conn.execute(
                'INSERT INTO users (username, password_hash) VALUES (?, ?)',
                (admin_username, generate_password_hash(admin_password))
            )

        # Create default project "Project One" and "FirstList" itemlist
        default_project = conn.execute(
            'SELECT id FROM projects WHERE name = ?', ('Project One',)
        ).fetchone()
        if not default_project:
            conn.execute(
                'INSERT INTO projects (name, description, status) VALUES (?, ?, ?)',
                ('Project One', 'Default project for organizing procurement', 'active')
            )
            default_project_id = conn.execute(
                'SELECT id FROM projects WHERE name = ?', ('Project One',)
            ).fetchone()['id']
            # Create "FirstList" itemlist for the default project
            conn.execute(
                'INSERT INTO item_lists (project_id, name, description) VALUES (?, ?, ?)',
                (default_project_id, 'FirstList', 'Default list for items')
            )
            # Move all items without a project into FirstList
            first_list_id = conn.execute(
                'SELECT id FROM item_lists WHERE project_id = ? AND name = ?',
                (default_project_id, 'FirstList')
            ).fetchone()['id']
            items = conn.execute('SELECT id FROM items').fetchall()
            for item in items:
                try:
                    conn.execute(
                        'INSERT INTO item_list_items (item_list_id, item_id) VALUES (?, ?)',
                        (first_list_id, item['id'])
                    )
                except Exception:
                    pass  # Item already in a list or duplicate


# ── Auth helpers ──────────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated


# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        with get_db() as conn:
            user = conn.execute('SELECT * FROM users WHERE username = ?', (username,)).fetchone()
        if user and check_password_hash(user['password_hash'], password):
            session['user_id'] = user['id']
            session['username'] = user['username']
            return redirect(request.args.get('next') or url_for('dashboard'))
        flash('Invalid username or password.', 'danger')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── Dashboard ─────────────────────────────────────────────────────────────────
@app.route('/dashboard')
@login_required
def dashboard():
    # Set default active project if not already set
    if 'active_project_id' not in session:
        with get_db() as conn:
            default_project = conn.execute(
                'SELECT id FROM projects WHERE name = ?', ('Project One',)
            ).fetchone()
            if default_project:
                session['active_project_id'] = default_project['id']

    with get_db() as conn:
        item_count     = conn.execute('SELECT COUNT(*) FROM items').fetchone()[0]
        supplier_count = conn.execute('SELECT COUNT(*) FROM suppliers').fetchone()[0]
        qr_count       = conn.execute('SELECT COUNT(*) FROM quote_requests').fetchone()[0]
        open_count     = conn.execute("SELECT COUNT(*) FROM quote_requests WHERE status='open'").fetchone()[0]
        project_count  = conn.execute('SELECT COUNT(*) FROM projects').fetchone()[0]
        recent = conn.execute('''
            SELECT sq.uploaded_at, s.name AS supplier_name, qr.title, qr.id AS qr_id
            FROM   supplier_quotes sq
            JOIN   suppliers      s  ON s.id  = sq.supplier_id
            JOIN   quote_requests qr ON qr.id = sq.quote_request_id
            ORDER  BY sq.uploaded_at DESC LIMIT 6
        ''').fetchall()
    return render_template('dashboard.html',
                           item_count=item_count, supplier_count=supplier_count,
                           qr_count=qr_count, open_count=open_count,
                           project_count=project_count, recent=recent)


# ── Item Lists (from Active Project) ──────────────────────────────────────────
@app.route('/item-lists')
@login_required
def item_lists_list():
    """Show item lists from the active project."""
    active_project_id = session.get('active_project_id')
    if not active_project_id:
        flash('No active project selected.', 'warning')
        return redirect(url_for('settings'))

    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (active_project_id,)).fetchone()
        if not project:
            abort(404)
        lists = conn.execute('''
            SELECT il.*,
                   COUNT(DISTINCT ili.item_id)  AS item_count,
                   COUNT(DISTINCT qr.id)         AS qr_count
            FROM   item_lists il
            LEFT JOIN item_list_items  ili ON ili.item_list_id = il.id
            LEFT JOIN quote_requests   qr  ON qr.item_list_id  = il.id
            WHERE  il.project_id = ?
            GROUP BY il.id
            ORDER BY il.created_at DESC
        ''', (active_project_id,)).fetchall()

    return render_template('item_lists/index.html', project=project, lists=lists)


# ── Settings (Project Management) ──────────────────────────────────────────────
@app.route('/settings')
@login_required
def settings():
    """Project management and settings."""
    with get_db() as conn:
        projects = conn.execute('''
            SELECT p.*,
                   COUNT(DISTINCT il.id)  AS list_count,
                   COUNT(DISTINCT qr.id)  AS qr_count
            FROM   projects p
            LEFT JOIN item_lists      il ON il.project_id   = p.id
            LEFT JOIN quote_requests  qr ON qr.item_list_id IN (
                          SELECT id FROM item_lists WHERE project_id = p.id)
            GROUP BY p.id
            ORDER BY p.created_at DESC
        ''').fetchall()

    return render_template('settings.html', projects=projects)


# ── Items ─────────────────────────────────────────────────────────────────────
@app.route('/items')
@login_required
def items_list():
    with get_db() as conn:
        items = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()

    # Group items by category
    from collections import defaultdict
    grouped = defaultdict(list)
    categories_set = set()

    for item in items:
        cat = item['category'] or 'Uncategorized'
        grouped[cat].append(item)
        categories_set.add(cat)

    # Sort categories with Uncategorized last
    categories = sorted(
        [c for c in categories_set if c != 'Uncategorized'],
        key=str.lower
    )
    if 'Uncategorized' in categories_set:
        categories.append('Uncategorized')

    return render_template('items/index.html', items=items, grouped_items=grouped, categories=categories)


@app.route('/items/<int:item_id>/qty', methods=['POST'])
@login_required
def items_update_qty(item_id):
    """Update item quantity inline."""
    try:
        qty = request.form.get('qty', '').strip()
        qty = float(qty) if qty else None
        with get_db() as conn:
            conn.execute('UPDATE items SET qty = ? WHERE id = ?', (qty, item_id))
        return jsonify({'success': True, 'qty': qty})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 400


def _read_df(raw, filename):
    buf = io.BytesIO(raw)
    return pd.read_csv(buf) if filename.lower().endswith('.csv') else pd.read_excel(buf)


def _guess(columns, candidates):
    """Return the first column name from candidates that exists (case-insensitive), else ''."""
    cl = [c.lower() for c in columns]
    for cand in candidates:
        if cand.lower() in cl:
            return columns[cl.index(cand.lower())]
    return ''


# ── Description formatting ────────────────────────────────────────────────────
def format_description(text):
    """
    Parse description text and return structured data for rendering.
    Detects headers (lines ending with :, markdown #), sections, and lists.
    Returns list of {'type': 'header'|'text'|'list'|'paragraph', 'content': str}
    """
    if not text or not text.strip():
        return []

    lines = text.split('\n')
    sections = []
    current_list = []

    for line in lines:
        stripped = line.strip()
        if not stripped:
            # Empty line
            if current_list:
                sections.append({'type': 'list', 'items': current_list})
                current_list = []
            sections.append({'type': 'spacing', 'content': ''})
            continue

        # Check for markdown headers
        if stripped.startswith('#'):
            if current_list:
                sections.append({'type': 'list', 'items': current_list})
                current_list = []
            level = len(stripped) - len(stripped.lstrip('#'))
            content = stripped.lstrip('#').strip()
            sections.append({'type': f'h{level}', 'content': content})

        # Check for header pattern (text ending with colon)
        elif stripped.endswith(':') and len(stripped) > 1:
            if current_list:
                sections.append({'type': 'list', 'items': current_list})
                current_list = []
            sections.append({'type': 'header', 'content': stripped[:-1]})

        # Check for bullet/list items
        elif stripped.startswith(('- ', '* ', '• ', '· ')):
            item = stripped[2:].strip()
            current_list.append(item)

        # Regular paragraph text
        else:
            if current_list:
                sections.append({'type': 'list', 'items': current_list})
                current_list = []
            sections.append({'type': 'paragraph', 'content': stripped})

    # Flush remaining list
    if current_list:
        sections.append({'type': 'list', 'items': current_list})

    return sections


# Step 1 – upload file, show mapping UI
@app.route('/items/import', methods=['GET', 'POST'])
@login_required
def items_import():
    if request.method == 'POST':
        file = request.files.get('file')
        if not file or file.filename == '':
            flash('No file selected.', 'danger')
            return redirect(request.url)
        if not allowed_file(file.filename):
            flash('Unsupported file type. Use .xlsx, .xls, or .csv', 'danger')
            return redirect(request.url)
        try:
            raw = file.read()
            df  = _read_df(raw, file.filename)
            columns = list(df.columns)
            if not columns:
                flash('File appears to be empty.', 'danger')
                return redirect(request.url)
            # Save temp file
            ts = datetime.now().strftime('%Y%m%d%H%M%S')
            tmp_name = secure_filename(f'tmp_items_{ts}_{file.filename}')
            with open(os.path.join(UPLOAD_FOLDER, tmp_name), 'wb') as fh:
                fh.write(raw)
            # Auto-guess mappings
            saved = load_mapping('items')
            guesses = {
                'name':        saved.get('name')        or _guess(columns, ['name', 'item_name', 'item', 'product', 'description']),
                'category':    saved.get('category')    or _guess(columns, ['category', 'cat', 'type', 'group']),
                'description': saved.get('description') or _guess(columns, ['description', 'desc', 'details', 'notes']),
                'qty':         saved.get('qty')         or _guess(columns, ['qty', 'quantity', 'amount', 'count']),
                'unit':        saved.get('unit')        or _guess(columns, ['unit', 'uom', 'measure']),
            }
            mode = request.form.get('mode', 'append')
            return render_template('items/map.html',
                                   columns=columns, tmp=tmp_name,
                                   guesses=guesses, mode=mode,
                                   preview=df.head(3).to_dict('records'))
        except Exception as e:
            flash(f'Error reading file: {e}', 'danger')
            return redirect(request.url)
    return render_template('items/import.html')


# Step 2 – apply mapping and import
@app.route('/items/import/do', methods=['POST'])
@login_required
def items_import_do():
    tmp      = request.form.get('tmp', '')
    mode     = request.form.get('mode', 'append')
    col_name = request.form.get('col_name', '').strip()
    col_cat  = request.form.get('col_category', '').strip()
    col_desc = request.form.get('col_description', '').strip()
    col_qty  = request.form.get('col_qty', '').strip()
    col_unit = request.form.get('col_unit', '').strip()

    if not tmp or not col_name:
        flash('Missing file or name column mapping.', 'danger')
        return redirect(url_for('items_import'))

    tmp_path = os.path.join(UPLOAD_FOLDER, secure_filename(tmp))
    if not os.path.exists(tmp_path):
        flash('Upload session expired. Please re-upload.', 'danger')
        return redirect(url_for('items_import'))

    try:
        with open(tmp_path, 'rb') as fh:
            raw = fh.read()
        df = _read_df(raw, tmp)

        def get_val(row, col):
            if not col or col not in df.columns:
                return None
            v = str(row.get(col, '') or '').strip()
            return None if v.lower() in ('', 'nan') else v

        with get_db() as conn:
            if mode == 'replace':
                conn.execute('DELETE FROM items')
            count = 0
            for _, row in df.iterrows():
                name = get_val(row, col_name)
                if not name:
                    continue
                qty = None
                if col_qty and col_qty in df.columns:
                    try:
                        qty_raw = row.get(col_qty)
                        if qty_raw is not None and str(qty_raw).lower() != 'nan':
                            qty = float(str(qty_raw).replace(',', ''))
                    except (ValueError, TypeError):
                        pass
                conn.execute(
                    'INSERT INTO items (name, category, description, qty, unit) VALUES (?,?,?,?,?)',
                    (name, get_val(row, col_cat), get_val(row, col_desc), qty, get_val(row, col_unit))
                )
                count += 1

        # Persist mapping for next time
        save_mapping('items', {
            'name': col_name, 'category': col_cat,
            'description': col_desc, 'qty': col_qty, 'unit': col_unit,
        })

        # Clean up temp file
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        flash(f'Imported {count} items successfully.', 'success')
        return redirect(url_for('items_list'))
    except Exception as e:
        flash(f'Error importing: {e}', 'danger')
        return redirect(url_for('items_import'))


@app.route('/items/edit/<int:item_id>', methods=['GET', 'POST'])
@login_required
def items_edit(item_id):
    with get_db() as conn:
        item = conn.execute('SELECT * FROM items WHERE id = ?', (item_id,)).fetchone()
        if not item:
            abort(404)
        if request.method == 'POST':
            conn.execute(
                'UPDATE items SET name=?,category=?,description=?,supplier_description=?,qty=?,unit=? WHERE id=?',
                (
                    request.form['name'].strip(),
                    request.form.get('category', '').strip() or None,
                    request.form.get('description', '').strip() or None,
                    request.form.get('supplier_description', '').strip() or None,
                    request.form.get('qty') or None,
                    request.form.get('unit', '').strip() or None,
                    item_id,
                )
            )
            flash('Item updated.', 'success')
            return redirect(url_for('items_list'))
    return render_template('items/edit.html', item=item)


@app.route('/items/<int:item_id>/generate-supplier-desc', methods=['POST'])
@login_required
def generate_supplier_desc(item_id):
    """Generate supplier description from internal description using Claude AI."""
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured'}), 400

    with get_db() as conn:
        item = conn.execute('SELECT description FROM items WHERE id = ?', (item_id,)).fetchone()
        if not item or not item['description']:
            return jsonify({'error': 'No internal description found'}), 400

    try:
        from anthropic import Anthropic
        client = Anthropic(api_key=api_key)
        message = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=300,
            messages=[
                {
                    'role': 'user',
                    'content': f"""Given this internal item description, generate a concise technical specification suitable for sending to a supplier. Focus only on specs, materials, and measurements. Remove any internal notes, budget info, or procurement context. Plain text, max 3 sentences.

Internal Description:
{item['description']}

Supplier-facing description:"""
                }
            ]
        )
        supplier_desc = message.content[0].text.strip()
        return jsonify({'supplier_description': supplier_desc})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/items/delete/<int:item_id>', methods=['POST'])
@login_required
def items_delete(item_id):
    with get_db() as conn:
        conn.execute('DELETE FROM items WHERE id = ?', (item_id,))
    flash('Item deleted.', 'success')
    return redirect(url_for('items_list'))


# ── Projects ───────────────────────────────────────────────────────────────────
@app.route('/projects')
@login_required
def projects_list():
    with get_db() as conn:
        rows = conn.execute('''
            SELECT p.*,
                   COUNT(DISTINCT il.id)  AS list_count,
                   COUNT(DISTINCT qr.id)  AS qr_count
            FROM   projects p
            LEFT JOIN item_lists      il ON il.project_id   = p.id
            LEFT JOIN quote_requests  qr ON qr.item_list_id IN (
                          SELECT id FROM item_lists WHERE project_id = p.id)
            GROUP BY p.id
            ORDER BY p.created_at DESC
        ''').fetchall()
    return render_template('projects/index.html', projects=rows)


@app.route('/projects/create', methods=['GET', 'POST'])
@login_required
def projects_create():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Project name is required.', 'danger')
            return redirect(request.url)
        with get_db() as conn:
            conn.execute(
                'INSERT INTO projects (name, description, status) VALUES (?,?,?)',
                (name,
                 request.form.get('description', '').strip() or None,
                 request.form.get('status', 'active'))
            )
        flash('Project created.', 'success')
        return redirect(url_for('projects_list'))
    return render_template('projects/form.html', project=None, action='Create')


@app.route('/projects/<int:project_id>')
@login_required
def projects_detail(project_id):
    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        if not project:
            abort(404)
        lists = conn.execute('''
            SELECT il.*,
                   COUNT(DISTINCT ili.item_id)  AS item_count,
                   COUNT(DISTINCT qr.id)         AS qr_count
            FROM   item_lists il
            LEFT JOIN item_list_items  ili ON ili.item_list_id = il.id
            LEFT JOIN quote_requests   qr  ON qr.item_list_id  = il.id
            WHERE  il.project_id = ?
            GROUP BY il.id
            ORDER BY il.created_at DESC
        ''', (project_id,)).fetchall()
    session['active_project_id'] = project_id
    return render_template('projects/detail.html', project=project, lists=lists)


@app.route('/projects/<int:project_id>/edit', methods=['GET', 'POST'])
@login_required
def projects_edit(project_id):
    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        if not project:
            abort(404)
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if not name:
                flash('Project name is required.', 'danger')
                return redirect(request.url)
            conn.execute(
                'UPDATE projects SET name=?, description=?, status=? WHERE id=?',
                (name,
                 request.form.get('description', '').strip() or None,
                 request.form.get('status', 'active'),
                 project_id)
            )
            flash('Project updated.', 'success')
            return redirect(url_for('projects_list'))
    return render_template('projects/form.html', project=project, action='Edit')


@app.route('/projects/<int:project_id>/delete', methods=['POST'])
@login_required
def projects_delete(project_id):
    with get_db() as conn:
        conn.execute('DELETE FROM projects WHERE id = ?', (project_id,))
    flash('Project deleted.', 'success')
    return redirect(url_for('projects_list'))


# ── Item Lists ─────────────────────────────────────────────────────────────────
@app.route('/projects/<int:project_id>/lists/create', methods=['GET', 'POST'])
@login_required
def item_lists_create(project_id):
    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        if not project:
            abort(404)
        items = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()

        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if not name:
                flash('List name is required.', 'danger')
                return render_template('projects/item_list_form.html',
                                       project=project, item_list=None,
                                       items=items, selected_ids=set(),
                                       action='Create')
            item_ids = request.form.getlist('item_ids')

            cur = conn.execute(
                'INSERT INTO item_lists (project_id, name, description) VALUES (?,?,?)',
                (project_id,
                 name,
                 request.form.get('description', '').strip() or None)
            )
            list_id = cur.lastrowid
            for iid in item_ids:
                conn.execute(
                    'INSERT INTO item_list_items (item_list_id, item_id) VALUES (?,?)',
                    (list_id, int(iid))
                )
            flash('Item list created.', 'success')
            return redirect(url_for('item_lists_detail',
                                    project_id=project_id, list_id=list_id))

    return render_template('projects/item_list_form.html',
                           project=project, item_list=None,
                           items=items, selected_ids=set(),
                           action='Create', project_id=project_id)


@app.route('/projects/<int:project_id>/lists/<int:list_id>')
@login_required
def item_lists_detail(project_id, list_id):
    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        item_list = conn.execute(
            'SELECT * FROM item_lists WHERE id = ? AND project_id = ?',
            (list_id, project_id)
        ).fetchone()
        if not project or not item_list:
            abort(404)
        items = conn.execute('''
            SELECT i.* FROM items i
            JOIN item_list_items ili ON ili.item_id = i.id
            WHERE ili.item_list_id = ?
            ORDER BY i.category, i.name
        ''', (list_id,)).fetchall()
        quote_requests = conn.execute('''
            SELECT qr.*,
                   COUNT(DISTINCT qri.item_id)     AS item_count,
                   COUNT(DISTINCT qrs.supplier_id) AS supplier_count
            FROM   quote_requests qr
            LEFT JOIN quote_request_items     qri ON qri.quote_request_id = qr.id
            LEFT JOIN quote_request_suppliers qrs ON qrs.quote_request_id = qr.id
            WHERE  qr.item_list_id = ?
            GROUP BY qr.id
            ORDER BY qr.created_at DESC
        ''', (list_id,)).fetchall()
    session['active_project_id'] = project_id
    return render_template('projects/item_list_detail.html',
                           project=project, item_list=item_list,
                           items=items, quote_requests=quote_requests,
                           project_id=project_id, list_id=list_id)


@app.route('/projects/<int:project_id>/lists/<int:list_id>/edit', methods=['GET', 'POST'])
@login_required
def item_lists_edit(project_id, list_id):
    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        item_list = conn.execute(
            'SELECT * FROM item_lists WHERE id = ? AND project_id = ?',
            (list_id, project_id)
        ).fetchone()
        if not project or not item_list:
            abort(404)
        items = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()
        selected_ids = set(
            row['item_id'] for row in conn.execute(
                'SELECT item_id FROM item_list_items WHERE item_list_id = ?', (list_id,)
            ).fetchall()
        )

        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if not name:
                flash('List name is required.', 'danger')
                return render_template('projects/item_list_form.html',
                                       project=project, item_list=item_list,
                                       items=items, selected_ids=selected_ids,
                                       action='Edit')
            conn.execute(
                'UPDATE item_lists SET name=?, description=? WHERE id=?',
                (name,
                 request.form.get('description', '').strip() or None,
                 list_id)
            )
            new_item_ids = set(int(i) for i in request.form.getlist('item_ids'))
            # Replace all items: delete then re-insert
            conn.execute('DELETE FROM item_list_items WHERE item_list_id = ?', (list_id,))
            for iid in new_item_ids:
                conn.execute(
                    'INSERT INTO item_list_items (item_list_id, item_id) VALUES (?,?)',
                    (list_id, iid)
                )
            flash('Item list updated.', 'success')
            return redirect(url_for('item_lists_detail',
                                    project_id=project_id, list_id=list_id))

    return render_template('projects/item_list_form.html',
                           project=project, item_list=item_list,
                           items=items, selected_ids=selected_ids,
                           action='Edit')


@app.route('/projects/<int:project_id>/lists/<int:list_id>/delete', methods=['POST'])
@login_required
def item_lists_delete(project_id, list_id):
    with get_db() as conn:
        conn.execute('DELETE FROM item_lists WHERE id = ? AND project_id = ?',
                     (list_id, project_id))
    flash('Item list deleted.', 'success')
    return redirect(url_for('projects_detail', project_id=project_id))


# ── Quote Request from Item List ───────────────────────────────────────────────
@app.route('/projects/<int:project_id>/lists/<int:list_id>/quotes/create',
           methods=['GET', 'POST'])
@login_required
def item_list_quotes_create(project_id, list_id):
    with get_db() as conn:
        project = conn.execute('SELECT * FROM projects WHERE id = ?', (project_id,)).fetchone()
        item_list = conn.execute(
            'SELECT * FROM item_lists WHERE id = ? AND project_id = ?',
            (list_id, project_id)
        ).fetchone()
        if not project or not item_list:
            abort(404)

        all_items = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()
        list_item_ids = set(
            row['item_id'] for row in conn.execute(
                'SELECT item_id FROM item_list_items WHERE item_list_id = ?', (list_id,)
            ).fetchall()
        )
        suppliers = conn.execute('SELECT * FROM suppliers ORDER BY name').fetchall()

        if request.method == 'POST':
            title        = request.form.get('title', '').strip()
            notes        = request.form.get('notes', '').strip()
            item_ids     = request.form.getlist('item_ids')
            supplier_ids = request.form.getlist('supplier_ids')

            errors = []
            if not title:        errors.append('Title is required.')
            if not item_ids:     errors.append('Select at least one item.')
            if not supplier_ids: errors.append('Select at least one supplier.')
            for e in errors:
                flash(e, 'danger')
            if errors:
                return render_template('projects/quote_create.html',
                                       project=project, item_list=item_list,
                                       items=all_items, list_item_ids=list_item_ids,
                                       suppliers=suppliers, project_id=project_id,
                                       list_id=list_id)

            cur = conn.execute(
                'INSERT INTO quote_requests (title, notes, item_list_id) VALUES (?,?,?)',
                (title, notes or None, list_id)
            )
            qr_id = cur.lastrowid

            for iid in item_ids:
                conn.execute(
                    'INSERT INTO quote_request_items (quote_request_id, item_id) VALUES (?,?)',
                    (qr_id, int(iid))
                )
            for sid in supplier_ids:
                conn.execute(
                    'INSERT INTO quote_request_suppliers (quote_request_id, supplier_id) VALUES (?,?)',
                    (qr_id, int(sid))
                )

            flash('Quote request created.', 'success')
            return redirect(url_for('quotes_detail', qr_id=qr_id))

    return render_template('projects/quote_create.html',
                           project=project, item_list=item_list,
                           items=all_items, list_item_ids=list_item_ids,
                           suppliers=suppliers, project_id=project_id,
                           list_id=list_id)


# ── Suppliers ─────────────────────────────────────────────────────────────────
@app.route('/suppliers')
@login_required
def suppliers_list():
    with get_db() as conn:
        suppliers = conn.execute('SELECT * FROM suppliers ORDER BY name').fetchall()
    return render_template('suppliers/index.html', suppliers=suppliers)


@app.route('/suppliers/add', methods=['GET', 'POST'])
@login_required
def suppliers_add():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not name:
            flash('Supplier name is required.', 'danger')
            return redirect(request.url)
        with get_db() as conn:
            conn.execute(
                'INSERT INTO suppliers (name, phone, address, remarks) VALUES (?,?,?,?)',
                (name,
                 request.form.get('phone', '').strip() or None,
                 request.form.get('address', '').strip() or None,
                 request.form.get('remarks', '').strip() or None)
            )
        flash('Supplier added.', 'success')
        return redirect(url_for('suppliers_list'))
    return render_template('suppliers/form.html', supplier=None, action='Add')


@app.route('/suppliers/edit/<int:supplier_id>', methods=['GET', 'POST'])
@login_required
def suppliers_edit(supplier_id):
    with get_db() as conn:
        supplier = conn.execute('SELECT * FROM suppliers WHERE id = ?', (supplier_id,)).fetchone()
        if not supplier:
            abort(404)
        if request.method == 'POST':
            name = request.form.get('name', '').strip()
            if not name:
                flash('Supplier name is required.', 'danger')
                return redirect(request.url)
            conn.execute(
                'UPDATE suppliers SET name=?,phone=?,address=?,remarks=? WHERE id=?',
                (name,
                 request.form.get('phone', '').strip() or None,
                 request.form.get('address', '').strip() or None,
                 request.form.get('remarks', '').strip() or None,
                 supplier_id)
            )
            flash('Supplier updated.', 'success')
            return redirect(url_for('suppliers_list'))
    return render_template('suppliers/form.html', supplier=supplier, action='Edit')


@app.route('/suppliers/delete/<int:supplier_id>', methods=['POST'])
@login_required
def suppliers_delete(supplier_id):
    with get_db() as conn:
        conn.execute('DELETE FROM suppliers WHERE id = ?', (supplier_id,))
    flash('Supplier deleted.', 'success')
    return redirect(url_for('suppliers_list'))


# ── Quote Requests ────────────────────────────────────────────────────────────
@app.route('/quotes')
@login_required
def quotes_list():
    with get_db() as conn:
        rows = conn.execute('''
            SELECT qr.*,
                   COUNT(DISTINCT qri.item_id)     AS item_count,
                   COUNT(DISTINCT qrs.supplier_id) AS supplier_count,
                   COUNT(DISTINCT sq.id)           AS received_count,
                   il.name  AS list_name,
                   p.name   AS project_name,
                   p.id     AS project_id,
                   il.id    AS list_id
            FROM   quote_requests qr
            LEFT JOIN quote_request_items     qri ON qri.quote_request_id = qr.id
            LEFT JOIN quote_request_suppliers qrs ON qrs.quote_request_id = qr.id
            LEFT JOIN supplier_quotes         sq  ON sq.quote_request_id  = qr.id
            LEFT JOIN item_lists              il  ON il.id  = qr.item_list_id
            LEFT JOIN projects                p   ON p.id   = il.project_id
            GROUP BY qr.id
            ORDER BY qr.created_at DESC
        ''').fetchall()
    return render_template('quotes/index.html', requests=rows)


@app.route('/quotes/create', methods=['GET', 'POST'])
@login_required
def quotes_create():
    with get_db() as conn:
        items     = conn.execute('SELECT * FROM items ORDER BY category, name').fetchall()
        suppliers = conn.execute('SELECT * FROM suppliers ORDER BY name').fetchall()

        if request.method == 'POST':
            title       = request.form.get('title', '').strip()
            notes       = request.form.get('notes', '').strip()
            item_ids    = request.form.getlist('item_ids')
            supplier_ids = request.form.getlist('supplier_ids')

            errors = []
            if not title:       errors.append('Title is required.')
            if not item_ids:    errors.append('Select at least one item.')
            if not supplier_ids: errors.append('Select at least one supplier.')
            for e in errors:
                flash(e, 'danger')
            if errors:
                return render_template('quotes/create.html', items=items, suppliers=suppliers)

            cur = conn.execute('INSERT INTO quote_requests (title, notes) VALUES (?,?)', (title, notes or None))
            qr_id = cur.lastrowid
            for iid in item_ids:
                conn.execute('INSERT INTO quote_request_items (quote_request_id, item_id) VALUES (?,?)', (qr_id, int(iid)))
            for sid in supplier_ids:
                conn.execute('INSERT INTO quote_request_suppliers (quote_request_id, supplier_id) VALUES (?,?)', (qr_id, int(sid)))

            flash('Quote request created.', 'success')
            return redirect(url_for('quotes_detail', qr_id=qr_id))

    return render_template('quotes/create.html', items=items, suppliers=suppliers)


@app.route('/quotes/<int:qr_id>')
@login_required
def quotes_detail(qr_id):
    with get_db() as conn:
        qr = conn.execute('SELECT * FROM quote_requests WHERE id = ?', (qr_id,)).fetchone()
        if not qr:
            abort(404)
        items = conn.execute('''
            SELECT i.* FROM items i
            JOIN quote_request_items qri ON qri.item_id = i.id
            WHERE qri.quote_request_id = ?
            ORDER BY i.category, i.name
        ''', (qr_id,)).fetchall()
        suppliers = conn.execute('''
            SELECT s.* FROM suppliers s
            JOIN quote_request_suppliers qrs ON qrs.supplier_id = s.id
            WHERE qrs.quote_request_id = ?
            ORDER BY s.name
        ''', (qr_id,)).fetchall()
        sq_map = {}
        for s in suppliers:
            sq = conn.execute('''
                SELECT * FROM supplier_quotes
                WHERE quote_request_id = ? AND supplier_id = ?
                ORDER BY uploaded_at DESC LIMIT 1
            ''', (qr_id, s['id'])).fetchone()
            sq_map[s['id']] = sq
    return render_template('quotes/detail.html', qr=qr, items=items, suppliers=suppliers, sq_map=sq_map)


# Step 1 – upload file, show column mapping UI
@app.route('/quotes/<int:qr_id>/upload/<int:supplier_id>', methods=['GET', 'POST'])
@login_required
def quotes_upload(qr_id, supplier_id):
    with get_db() as conn:
        qr       = conn.execute('SELECT * FROM quote_requests WHERE id = ?', (qr_id,)).fetchone()
        supplier = conn.execute('SELECT * FROM suppliers WHERE id = ?', (supplier_id,)).fetchone()
        if not qr or not supplier:
            abort(404)
        items = conn.execute('''
            SELECT i.* FROM items i
            JOIN quote_request_items qri ON qri.item_id = i.id
            WHERE qri.quote_request_id = ?
            ORDER BY i.category, i.name
        ''', (qr_id,)).fetchall()

        if request.method == 'POST':
            notes = request.form.get('notes', '').strip()
            file  = request.files.get('file')

            if file and file.filename:
                if not allowed_file(file.filename):
                    flash('Unsupported file type. Use .xlsx, .xls, or .csv', 'danger')
                    return redirect(request.url)
                try:
                    raw     = file.read()
                    df      = _read_df(raw, file.filename)
                    columns = list(df.columns)
                    # Save temp file
                    ts       = datetime.now().strftime('%Y%m%d%H%M%S')
                    tmp_name = secure_filename(f'tmp_q{qr_id}_s{supplier_id}_{ts}_{file.filename}')
                    with open(os.path.join(UPLOAD_FOLDER, tmp_name), 'wb') as fh:
                        fh.write(raw)
                    saved = load_mapping('quotes')
                    guesses = {
                        'item_name': saved.get('item_name') or _guess(columns, ['item_name', 'name', 'item', 'product', 'description', 'material']),
                        'price':     saved.get('price')     or _guess(columns, ['price', 'unit_price', 'quoted_price', 'rate', 'amount', 'cost']),
                        'notes':     saved.get('notes')     or _guess(columns, ['notes', 'remarks', 'comment']),
                    }
                    return render_template('quotes/map.html',
                                           qr=qr, supplier=supplier, items=items,
                                           columns=columns, tmp=tmp_name, notes=notes,
                                           guesses=guesses,
                                           preview=df.head(3).to_dict('records'))
                except Exception as e:
                    flash(f'Error reading file: {e}', 'danger')
                    return redirect(request.url)

            # No file — manual-only entry path
            flash('Please upload a file or use the manual entry fields below.', 'warning')

    return render_template('quotes/upload.html', qr=qr, supplier=supplier, items=items)


# Step 2a – apply column mapping from uploaded file (+ optional manual overrides)
@app.route('/quotes/<int:qr_id>/upload/<int:supplier_id>/do', methods=['POST'])
@login_required
def quotes_upload_do(qr_id, supplier_id):
    with get_db() as conn:
        qr       = conn.execute('SELECT * FROM quote_requests WHERE id = ?', (qr_id,)).fetchone()
        supplier = conn.execute('SELECT * FROM suppliers WHERE id = ?', (supplier_id,)).fetchone()
        if not qr or not supplier:
            abort(404)
        items = conn.execute('''
            SELECT i.* FROM items i
            JOIN quote_request_items qri ON qri.item_id = i.id
            WHERE qri.quote_request_id = ?
            ORDER BY i.category, i.name
        ''', (qr_id,)).fetchall()

        tmp        = request.form.get('tmp', '')
        notes      = request.form.get('notes', '').strip()
        col_name   = request.form.get('col_item_name', '').strip()
        col_price  = request.form.get('col_price', '').strip()
        col_notes  = request.form.get('col_notes', '').strip()
        prices_map = {}

        if tmp:
            tmp_path = os.path.join(UPLOAD_FOLDER, secure_filename(tmp))
            if not os.path.exists(tmp_path):
                flash('Upload session expired. Please re-upload.', 'danger')
                return redirect(url_for('quotes_upload', qr_id=qr_id, supplier_id=supplier_id))
            try:
                with open(tmp_path, 'rb') as fh:
                    raw = fh.read()
                df         = _read_df(raw, tmp)
                name_to_id = {i['name'].lower().strip(): i['id'] for i in items}

                if col_name and col_price:
                    for _, row in df.iterrows():
                        key = str(row.get(col_name, '')).lower().strip()
                        if key in name_to_id and row.get(col_price) is not None:
                            try:
                                prices_map[name_to_id[key]] = float(str(row[col_price]).replace(',', ''))
                            except (ValueError, TypeError):
                                pass

                # Rename to permanent file
                ts    = datetime.now().strftime('%Y%m%d%H%M%S')
                fname = secure_filename(f'q{qr_id}_s{supplier_id}_{ts}_{tmp.split("_", 4)[-1]}')
                os.rename(tmp_path, os.path.join(UPLOAD_FOLDER, fname))
                file_path = fname
            except Exception as e:
                flash(f'Error processing file: {e}', 'danger')
                return redirect(url_for('quotes_upload', qr_id=qr_id, supplier_id=supplier_id))
        else:
            file_path = None

        # Manual price overrides always win
        for item in items:
            val = request.form.get(f'price_{item["id"]}', '').strip()
            if val:
                try:
                    prices_map[item['id']] = float(val.replace(',', ''))
                except ValueError:
                    pass

        if not prices_map:
            flash('No prices found. Please map columns correctly or enter prices manually.', 'warning')
            return redirect(url_for('quotes_upload', qr_id=qr_id, supplier_id=supplier_id))

        cur = conn.execute(
            'INSERT INTO supplier_quotes (quote_request_id, supplier_id, file_path, notes) VALUES (?,?,?,?)',
            (qr_id, supplier_id, file_path, notes or None)
        )
        sq_id = cur.lastrowid
        for item_id, price in prices_map.items():
            conn.execute(
                'INSERT INTO quote_prices (supplier_quote_id, item_id, price) VALUES (?,?,?)',
                (sq_id, item_id, price)
            )

    # Persist mapping for next time
    if tmp:
        save_mapping('quotes', {'item_name': col_name, 'price': col_price, 'notes': col_notes})

    flash(f'Quote from {supplier["name"]} saved with {len(prices_map)} prices.', 'success')
    return redirect(url_for('quotes_detail', qr_id=qr_id))


@app.route('/quotes/<int:qr_id>/comparison')
@login_required
def quotes_comparison(qr_id):
    with get_db() as conn:
        qr = conn.execute('SELECT * FROM quote_requests WHERE id = ?', (qr_id,)).fetchone()
        if not qr:
            abort(404)
        items = conn.execute('''
            SELECT i.* FROM items i
            JOIN quote_request_items qri ON qri.item_id = i.id
            WHERE qri.quote_request_id = ?
            ORDER BY i.category, i.name
        ''', (qr_id,)).fetchall()
        suppliers = conn.execute('''
            SELECT s.* FROM suppliers s
            JOIN quote_request_suppliers qrs ON qrs.supplier_id = s.id
            WHERE qrs.quote_request_id = ?
            ORDER BY s.name
        ''', (qr_id,)).fetchall()

        # price_matrix[item_id][supplier_id] = price
        price_matrix = {item['id']: {} for item in items}
        for s in suppliers:
            sq = conn.execute('''
                SELECT id FROM supplier_quotes
                WHERE quote_request_id = ? AND supplier_id = ?
                ORDER BY uploaded_at DESC LIMIT 1
            ''', (qr_id, s['id'])).fetchone()
            if sq:
                for p in conn.execute('SELECT item_id, price FROM quote_prices WHERE supplier_quote_id = ?', (sq['id'],)).fetchall():
                    if p['item_id'] in price_matrix:
                        price_matrix[p['item_id']][s['id']] = p['price']

    return render_template('quotes/comparison.html',
                           qr=qr, items=items, suppliers=suppliers, price_matrix=price_matrix)


@app.route('/quotes/<int:qr_id>/toggle-status', methods=['POST'])
@login_required
def quotes_toggle_status(qr_id):
    with get_db() as conn:
        qr = conn.execute('SELECT status FROM quote_requests WHERE id = ?', (qr_id,)).fetchone()
        if qr:
            new = 'closed' if qr['status'] == 'open' else 'open'
            conn.execute('UPDATE quote_requests SET status = ? WHERE id = ?', (new, qr_id))
    return redirect(url_for('quotes_detail', qr_id=qr_id))


@app.route('/quotes/<int:qr_id>/delete', methods=['POST'])
@login_required
def quotes_delete(qr_id):
    with get_db() as conn:
        conn.execute('DELETE FROM quote_requests WHERE id = ?', (qr_id,))
    flash('Quote request deleted.', 'success')
    return redirect(url_for('quotes_list'))


@app.route('/quotes/<int:qr_id>/export', methods=['GET'])
@login_required
def quotes_export(qr_id):
    """Export quote request as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        with get_db() as conn:
            qr = conn.execute('SELECT * FROM quote_requests WHERE id = ?', (qr_id,)).fetchone()
            if not qr:
                abort(404)

            items = conn.execute('''
                SELECT i.* FROM items i
                JOIN quote_request_items qri ON qri.item_id = i.id
                WHERE qri.quote_request_id = ?
                ORDER BY i.category, i.name
            ''', (qr_id,)).fetchall()

            suppliers = conn.execute('''
                SELECT s.* FROM suppliers s
                JOIN quote_request_suppliers qrs ON qrs.supplier_id = s.id
                WHERE qrs.quote_request_id = ?
                ORDER BY s.name
            ''', (qr_id,)).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Items'

        # Header row styling
        header_fill = PatternFill(start_color='4F8EF7', end_color='4F8EF7', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')

        # Sheet 1: Items
        headers = ['Item Name', 'Category', 'Quantity', 'Unit', 'Supplier Description']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font

        for row_idx, item in enumerate(items, 2):
            # Handle supplier_description safely - it may not exist in older DBs
            supplier_desc = ''
            try:
                supplier_desc = item['supplier_description'] or item['description'] or ''
            except (KeyError, TypeError):
                supplier_desc = item.get('description') or ''

            ws.cell(row=row_idx, column=1, value=item['name'])
            ws.cell(row=row_idx, column=2, value=item['category'] or '')
            ws.cell(row=row_idx, column=3, value=item['qty'] if item['qty'] is not None else '')
            ws.cell(row=row_idx, column=4, value=item['unit'] or '')
            ws.cell(row=row_idx, column=5, value=supplier_desc)

        # Auto-width columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 40

        # Freeze top row
        ws.freeze_panes = 'A2'

        # Sheet 2: Details
        ws2 = wb.create_sheet('Details')
        ws2.cell(row=1, column=1, value='Quote Request Details').font = Font(bold=True, size=12)
        ws2.cell(row=3, column=1, value='Title:').font = Font(bold=True)
        ws2.cell(row=3, column=2, value=qr['title'])
        ws2.cell(row=4, column=1, value='Created:').font = Font(bold=True)
        ws2.cell(row=4, column=2, value=qr['created_at'])
        ws2.cell(row=5, column=1, value='Status:').font = Font(bold=True)
        ws2.cell(row=5, column=2, value=qr['status'].capitalize())
        if qr['notes']:
            ws2.cell(row=6, column=1, value='Notes:').font = Font(bold=True)
            ws2.cell(row=6, column=2, value=qr['notes'])

        ws2.cell(row=8, column=1, value='Suppliers:').font = Font(bold=True)
        for row_idx, supplier in enumerate(suppliers, 9):
            ws2.cell(row=row_idx, column=1, value=supplier['name'])
            ws2.cell(row=row_idx, column=2, value=supplier['phone'] or '')

        ws2.column_dimensions['A'].width = 20
        ws2.column_dimensions['B'].width = 35

        # Return as file
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        filename = f"quote-{qr['title'][:20]}-{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(
            buf,
            download_name=filename,
            as_attachment=True,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        flash(f'Error exporting quote: {str(e)}', 'danger')
        return redirect(request.referrer or url_for('quotes_list'))


# ── Global Price Comparison ───────────────────────────────────────────────────
@app.route('/comparison')
@login_required
def comparison_global():
    """Show global price comparison across all quote requests and suppliers."""
    with get_db() as conn:
        # Only items that have at least one price
        items = conn.execute('''
            SELECT DISTINCT i.id, i.name, i.category, i.qty, i.unit
            FROM items i
            JOIN quote_prices qp ON qp.item_id = i.id
            ORDER BY i.category, i.name
        ''').fetchall()

        # Only suppliers that have actually submitted prices
        suppliers = conn.execute('''
            SELECT DISTINCT s.id, s.name
            FROM suppliers s
            JOIN supplier_quotes sq ON sq.supplier_id = s.id
            JOIN quote_prices qp ON qp.supplier_quote_id = sq.id
            ORDER BY s.name
        ''').fetchall()

        # Global best price per (item, supplier) across all quote requests
        price_matrix = {item['id']: {} for item in items}
        raw_prices = conn.execute('''
            SELECT qp.item_id, sq.supplier_id, MIN(qp.price) AS best_price
            FROM quote_prices qp
            JOIN supplier_quotes sq ON sq.id = qp.supplier_quote_id
            WHERE qp.price IS NOT NULL
            GROUP BY qp.item_id, sq.supplier_id
        ''').fetchall()

        for row in raw_prices:
            if row['item_id'] in price_matrix:
                price_matrix[row['item_id']][row['supplier_id']] = row['best_price']

    return render_template('comparison.html',
                           items=items,
                           suppliers=suppliers,
                           price_matrix=price_matrix)


# ── Uploaded file download ────────────────────────────────────────────────────
@app.route('/uploads/<path:filename>')
@login_required
def uploaded_file(filename):
    return send_from_directory(UPLOAD_FOLDER, filename, as_attachment=True)


# ── Bootstrap ─────────────────────────────────────────────────────────────────
init_db()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
